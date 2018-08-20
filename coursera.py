import requests
import random
import argparse
from argparse import ArgumentTypeError
from bs4 import BeautifulSoup
from requests import ReadTimeout
from requests import HTTPError
from requests import ConnectionError
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from os import access
from os import W_OK
from os import path
from os import getcwd
from collections import OrderedDict
from course_parser import (get_user_score, get_date_start, get_course_language,
                           get_course_name, get_weeks_count)


def check_file_path(file_path):
    if not access(path.dirname(file_path), W_OK):
        raise ArgumentTypeError(
            "You don't have permissions to '{}' , "
            "or directory doesn't exist!".format(file_path)
        )
    elif path.isdir(file_path):
        raise ArgumentTypeError(
            "The '{}' is not a file!".format(file_path)
        )
    else:
        return file_path


def get_args():
    script_usage = "python coursera.py  <path to output file>"
    parser = argparse.ArgumentParser(
        description="How to run coursera.py:",
        usage=script_usage
    )
    parser.add_argument(
        "output_file_path",
        type=check_file_path,
        nargs='?',
        default=path.join(getcwd(), "coursera.xlsx"),
        help="Specify the path to output file"
    )
    args = parser.parse_args()
    return args


def fetch_content_from_url(url):
    try:
        response = requests.get(url)
        return response.content
    except (ConnectionError, HTTPError, ReadTimeout):
        return None


def get_random_courses_urls_list(sitemap_content, courses_count=20):
    soup = BeautifulSoup(sitemap_content, "html.parser")
    urls = [element.text for element in soup.findAll("loc")]
    courses_list = random.sample(urls, courses_count)
    return courses_list


def get_course_info(course_content):
    if not course_content:
        return None
    course = BeautifulSoup(course_content, "lxml")
    course_info = OrderedDict()
    course_info["course_name"] = get_course_name(course)
    course_info["language"] = get_course_language(course)
    course_info["date_start"] = get_date_start(course)
    course_info["weeks_count"] = get_weeks_count(course)
    course_info["user_score"] = get_user_score(course)
    return course_info


def make_xlsx_template():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Courses List"
    fill = PatternFill(
        fill_type="solid",
        start_color="c1c1c1",
        end_color="c2c2c2"
    )
    head_row = [
        "URL",
        "COURSE_NAME",
        "LANGUAGE",
        "NEAREST_DATE",
        "WEEKS_COUNT",
        "USER_SCORE"
    ]
    worksheet.append(head_row)
    for cell in ["A1", "B1", "C1", "D1", "E1", "F1"]:
        worksheet[cell].fill = fill
    return workbook, worksheet


def output_courses_info_to_xlsx(courses_data, workbook, worksheet):
    for url, course_info in courses_data.items():
        try:
            next_row = [
                url,
                course_info["course_name"],
                course_info["language"],
                course_info["date_start"],
                course_info["weeks_count"],
                course_info["user_score"]
            ]
        except TypeError:
            next_row = [url]
        worksheet.append(next_row)
    return workbook


def save_output_courses_to_xlsx(wb, filepath):
    wb.save(filepath)


if __name__ == "__main__":
    args = get_args()
    output_file_path = args.output_file_path
    sitemap_content = fetch_content_from_url(
        "https://www.coursera.org/sitemap~www~courses.xml")
    if not sitemap_content:
        exit("The www.coursera.org is unavailable!")
    courses_urls_list = get_random_courses_urls_list(
        sitemap_content,
        courses_count=3
    )
    courses_info = {}
    for course_url in courses_urls_list:
        course_content = fetch_content_from_url(course_url)
        courses_info[course_url] = get_course_info(course_content)
    workbook, worksheet = make_xlsx_template()
    xlsx_courses_info = output_courses_info_to_xlsx(
        courses_info,
        workbook,
        worksheet
    )
    save_output_courses_to_xlsx(
        xlsx_courses_info,
        filepath=output_file_path
    )
    print("Saved successfully to the {}.".format(output_file_path))
